#!/usr/bin/env python3
#v2
"""
mastercard_iso8583_parser.py
=============================
Mastercard Dual Message Authorization System (DMS) ISO 8583 message parser.

Modeled directly on visa_iso8583_parser_v2.py's architecture (same compact
output convention: 'FieldNum)value' joined by 'Ý', subfields joined by
'³'), but the wire format underneath is meaningfully different from the
Visa parser and every difference below was confirmed by decoding real
sample traffic byte-for-byte against known-correct parsed output, not by
reading the spec alone.

CONFIRMED STRUCTURAL DIFFERENCES FROM THE VISA PARSER:
  - No proprietary binary header. The message is just MTI + bitmap(s) +
    fields, all the way through. MTI is 4 EBCDIC DIGIT CHARACTERS (e.g.
    hex F0F1F0F0 -> "0100"), not BCD nibbles like Visa's header/MTI.
  - Bitmaps are standard: primary bit 1 = "secondary bitmap follows"
    flag (fields 2-64 map directly to primary bits 2-64). The secondary
    bitmap has NO analogous flag-reinterpretation trick the way the Visa
    parser needed - secondary bit N maps directly to field 64+N (e.g.
    secondary bit 58 = field 122), confirmed against a real F122 sample.
    No tertiary bitmap has been observed or is coded for.
  - The overwhelming majority of fields are PLAIN EBCDIC DIGIT/TEXT
    CHARACTERS, not packed BCD. Masked digits and Track 2 field
    separators already arrive as literal EBCDIC characters ('*' and '=')
    - no nibble-level mask/separator translation is needed the way the
      Visa parser needed it for F2/F35.
  - Variable-length fields use ASCII/EBCDIC-digit length prefixes of
    varying width depending on the field: LLVAR (2-digit prefix) for
    F2/F32/F33/F35, and LLLVAR (3-digit prefix) for F48/F55/F61/F63/F90's
    neighbors F122. This was confirmed field-by-field against samples,
    not assumed from a single convention.
  - F48 (Additional Data - Private) begins with ONE UNTAGGED single
    EBCDIC character - this is DE 48's "Transaction Category Code" that
    precedes the subelement list per spec (not a normal TLV subelement)
    - followed by a repeating chain of tag(2 EBCDIC digits) +
      len(2 EBCDIC digits) + value. Confirmed against two independent
      samples (request and response), including a case with 6 chained
      subelements.
  - F55 (ICC/EMV data) is a 3-digit LLLVAR length prefix followed by a
    STANDARD BINARY BER-TLV chain (real binary tag/length bytes, e.g.
    0x9F 0x26 0x08 ..., NOT EBCDIC-encoded like the rest of the
    message). Reuses the same generic BER-TLV reader as the Visa parser
    since that code already operates on raw bytes.
  - F122 is a 3-digit LLLVAR outer length, then a FLAT (non-recursive)
    repeating chain of tag(3 digits) + len(3 digits) + value. Confirmed
    against a sample where the single tag "001"'s value itself contained
    text that superficially looks like further tag/len/value triples
    (e.g. "002003DHA003010AROGGA.COM...") - the reference tooling this
    parser is matched against does NOT decompose that further, so this
    parser deliberately does not either. Do not "fix" this into a
    recursive/nested parse without new sample evidence showing the
    reference output actually expects that.
  - F43 (Card Acceptor Name/Location) is fixed 40 EBCDIC chars, split
    into 5 positional subfields with widths [22, 1, 13, 1, 3] (Name,
    separator, City, separator, Country) - confirmed against two
    independent samples with different merchant names/cities.
  - F90 (Original Data Elements) is fixed 42 EBCDIC chars, split into
    5 positional subfields with widths [4, 6, 10, 11, 11] (original MTI,
    original STAN, original transmission date/time, original amount,
    and a 5th 11-digit block whose meaning is NOT yet confirmed - it has
    only been observed as all-zero so far).
  - F52 (PIN block): fixed 8 bytes, always rendered as 16 '*' (same
    masking convention as the Visa parser), confirmed against a sample
    carrying an all-0xFF PIN block.

KNOWN OPEN ITEMS (flag for more sample data before trusting fully):
  - Only primary + secondary bitmap confirmed (fields 2-128). No
    tertiary bitmap support - unconfirmed whether Mastercard DMS traffic
    ever uses one for this message family.
  - F61's internal subfield structure (POS Data) is NOT yet broken out -
    rendered as a single raw string, same as the Visa parser did for its
    own F61 before confirmation. Per spec DE 61 has multiple positional
    subfields; needs a sample where subfield boundaries can be verified.
  - F122's only confirmed subelement tag so far is "001". Any other tag
    is rendered with a generic "Subelement {tag}" label.
  - F48 subelement tag meanings are confirmed by NUMBER (from the spec's
    "List of DE 48 subelements") for tags 32/37/42/43/61/63/66/71/87/92,
    but their internal VALUE meanings (beyond raw pass-through) are not
    yet transcribed into mastercard_field_details.py except where a
    sample made the meaning unambiguous.
  - F55 EMV tag value rendering (hex vs. decimal) is only confirmed for
    the tags that appeared in the sample traffic; anything else defaults
    to hex passthrough, mirroring the Visa parser's conservative default.
  - Advice (0120/0130), sign-on/echo (0800/0810), and file update
    (0302/0312) message families have not been sampled yet - only
    0100/0110 (authorization) and 0400/0410 (reversal) are confirmed.
"""

import re
import sys
from datetime import datetime

try:
    from mastercard_field_details import get_value_detail, format_field_report
except ImportError:
    def get_value_detail(label, value):
        return "mastercard_field_details.py not found - per-value detail lookup unavailable."

    def format_field_report(rows):
        return "mastercard_field_details.py not found - per-value detail lookup unavailable."

FIELD_SEP = "\u00dd"      # Ý  - separator between fields
SUBFIELD_SEP = "\u00b3"   # ³  - separator between repeated subfield groups

# ---------------------------------------------------------------------------
# Low level helpers
# ---------------------------------------------------------------------------

def clean_hex(raw: str) -> bytes:
    hex_str = re.sub(r"\s+", "", raw)
    if len(hex_str) % 2:
        raise ValueError(
            f"Hex string has an odd number of characters ({len(hex_str)})"
        )
    return bytes.fromhex(hex_str)


def decode_ebcdic(data: bytes) -> str:
    try:
        return data.decode("cp037")
    except Exception:
        return data.hex().upper()


def is_fully_printable_ebcdic(data: bytes) -> bool:
    try:
        s = data.decode("cp037")
    except Exception:
        return False
    return all(c.isprintable() for c in s)


def to_hex(data: bytes) -> str:
    return data.hex().upper()


def bin_to_int(data: bytes) -> int:
    return int.from_bytes(data, "big")


# ---------------------------------------------------------------------------
# BER-TLV decoding (F55 - reused verbatim-style from the Visa parser since
# it already operates on raw bytes; Mastercard's F55 uses real binary tags)
# ---------------------------------------------------------------------------

def read_tag(data: bytes, pos: int):
    first = data[pos]
    tag_bytes = [first]
    pos += 1
    if (first & 0x1F) == 0x1F:  # low 5 bits all 1 -> multi-byte tag
        while True:
            b = data[pos]
            tag_bytes.append(b)
            pos += 1
            if not (b & 0x80):
                break
    return bytes(tag_bytes).hex(), pos


def read_length(data: bytes, pos: int):
    first = data[pos]
    pos += 1
    if first & 0x80:  # long form
        num_bytes = first & 0x7F
        length = int.from_bytes(data[pos:pos + num_bytes], "big")
        pos += num_bytes
    else:  # short form
        length = first & 0x7F
    return length, pos


def render_hex(tag: str, value: bytes) -> str:
    return to_hex(value)


def render_int_or_empty(tag: str, value: bytes) -> str:
    n = int.from_bytes(value, "big")
    return "" if n == 0 else str(n)


# Tags confirmed to render as plain decimal integers rather than hex,
# verified against live sample traffic:
#   9F27 (CID)                  single byte 0x80          -> "128"
#   82   (AIP)                  bytes 0x3900              -> "14592"
#   9F33 (Terminal Capabilities) bytes 0x604020            -> "6307872"
# Everything else defaults to hex passthrough until confirmed otherwise.
F55_DECIMAL_TAGS = {"9f27", "82", "9f33"}


def render_f55_value(tag: str, value: bytes) -> str:
    if tag in F55_DECIMAL_TAGS:
        return render_int_or_empty(tag, value)
    return render_hex(tag, value)


def decode_plain_tlv(data: bytes, value_renderer=render_hex) -> dict:
    """Chained Tag-Length-Value elements, no dataset wrapper (used for F55)."""
    out = {}
    pos = 0
    n = len(data)
    while pos < n:
        tag, pos = read_tag(data, pos)
        length, pos = read_length(data, pos)
        value = data[pos:pos + length]
        pos += length
        out[tag] = value_renderer(tag, value)
    return out


# ---------------------------------------------------------------------------
# Header / MTI parsing
# ---------------------------------------------------------------------------

def parse_header(data: bytes):
    """Mastercard DMS traffic (confirmed): no proprietary header bytes
    before the MTI - the message starts directly with 4 EBCDIC MTI digit
    characters, unlike the Visa parser's binary header block."""
    mti = decode_ebcdic(data[0:4])
    return {"MTI": mti}, 4


# ---------------------------------------------------------------------------
# Bitmap parsing
# ---------------------------------------------------------------------------

def parse_bitmap_block(data: bytes, pos: int, field_offset: int, has_flag: bool):
    """
    Parse one 8-byte bitmap block.
    - Primary block (has_flag=True): bit 1 is the "secondary bitmap
      follows" FLAG (not a field); bits 2-64 map to fields
      (field_offset+1)..(field_offset+63).
    - Secondary block (has_flag=False): CONFIRMED direct mapping - bit N
      maps straight to field (field_offset+N) with no flag/offset trick
      (e.g. secondary bit 58 -> field 122, confirmed against a live
      sample). No tertiary bitmap support - unconfirmed for this message
      family.
    """
    block = data[pos:pos + 8]
    pos += 8
    bits = []
    for byte in block:
        for i in range(8):
            bits.append((byte >> (7 - i)) & 1)
    present = {}
    flag = 0
    if has_flag:
        flag = bits[0]
        for i in range(1, 64):
            if bits[i]:
                present[field_offset + i] = True
    else:
        for i in range(0, 64):
            if bits[i]:
                present[field_offset + i + 1] = True
    return present, flag, pos


def parse_all_bitmaps(data: bytes, pos: int):
    present = {}
    primary, flag2, pos = parse_bitmap_block(data, pos, field_offset=1, has_flag=True)   # fields 2-64
    present.update(primary)
    if flag2:
        secondary, _flag, pos = parse_bitmap_block(data, pos, field_offset=64, has_flag=False)  # fields 65-128
        present.update(secondary)
    return present, pos


# ---------------------------------------------------------------------------
# Field definitions
# ---------------------------------------------------------------------------
# type: 'an'      fixed-length plain EBCDIC characters (digits or text)
#       'llvar'   2-EBCDIC-digit length prefix + EBCDIC value
#       'lllvar'  3-EBCDIC-digit length prefix + EBCDIC value (raw pass-through
#                 unless a 'special' handler is registered for the field)
#       'pin'     fixed 8 bytes, always fully masked
#       'special' -> custom handler keyed by field number in FIELD_DEFS

FIELD_DEFS = {
    2:  {"type": "llvar"},                                      # PAN
    3:  {"type": "special_3"},                                  # Processing code (3 x 2-digit subfields)
    4:  {"type": "an", "len": 12},                              # Amount, transaction
    5:  {"type": "an", "len": 12},                              # Amount, reconciliation
    6:  {"type": "an", "len": 12},                              # Amount, cardholder billing
    7:  {"type": "an", "len": 10},                              # Transmission date/time
    9:  {"type": "an", "len": 8},                               # Conversion rate, reconciliation
    10: {"type": "an", "len": 8},                               # Conversion rate, cardholder billing
    11: {"type": "an", "len": 6},                               # STAN
    12: {"type": "an", "len": 6},                               # Time, local
    13: {"type": "an", "len": 4},                               # Date, local
    14: {"type": "an", "len": 4},                               # Date, expiration
    15: {"type": "an", "len": 4},                               # Date, settlement
    16: {"type": "an", "len": 4},                               # Date, conversion
    18: {"type": "an", "len": 4},                               # Merchant Category Code (MCC)
    19: {"type": "an", "len": 3},                               # Acquiring inst country code
    22: {"type": "an", "len": 3},                               # POS entry mode
    23: {"type": "an", "len": 3},                               # Card sequence number
    26: {"type": "an", "len": 2},                               # POS PIN Capture Code
    28: {"type": "special_28"},                                 # Amount, Transaction Fee
    32: {"type": "llvar"},                                      # Acquiring inst ID code
    33: {"type": "llvar"},                                      # Forwarding inst ID code
    35: {"type": "llvar"},                                      # Track 2 data (mask/sep chars already literal EBCDIC)
    37: {"type": "an", "len": 12},                              # RRN
    38: {"type": "an", "len": 6},                               # Auth ID response
    39: {"type": "an", "len": 2},                               # Response code
    41: {"type": "an", "len": 8},                               # Card acceptor terminal ID
    42: {"type": "an", "len": 15},                              # Card acceptor ID code
    43: {"type": "special_43"},                                 # Card acceptor name/location (positional)
    44: {"type": "llvar"},                                      # Additional Response Data
    48: {"type": "special_48"},                                 # Additional Data - Private
    49: {"type": "an", "len": 3},                               # Currency code, transaction
    50: {"type": "an", "len": 3},                               # Currency code, reconciliation
    51: {"type": "an", "len": 3},                               # Currency code, cardholder billing
    52: {"type": "pin"},                                        # PIN data
    54: {"type": "special_54"},                                 # Additional Amounts
    55: {"type": "special_55"},                                 # ICC / EMV data
    56: {"type": "lllvar"},                                     # Payment Account Data (PAR)
    61: {"type": "lllvar"},                                     # POS Data (raw pass-through - unconfirmed subfields)
    62: {"type": "lllvar"},                                     # Intermediate Network Facility (INF) Data
    63: {"type": "lllvar"},                                     # Private/network reference (raw pass-through)
    67: {"type": "an", "len": 2},                               # Extended Payment Code
    90: {"type": "special_90"},                                 # Original Data Elements (positional)
    94: {"type": "special_94"},                                 # Service Indicator
    95: {"type": "special_95"},                                 # Replacement Amounts
    96: {"type": "an", "len": 8},                               # Message Security Code
    102: {"type": "llvar"},                                     # Account ID 1
    103: {"type": "llvar"},                                     # Account ID 2
    104: {"type": "special_104"},                               # Digital Payment Data
    105: {"type": "special_105"},                               # Multi-Use Transaction Identification Data
    106: {"type": "special_106"},                               # Fleet Service Data
    108: {"type": "special_108"},                               # Additional Transaction Reference Data
    110: {"type": "special_110"},                               # Additional Data-2 / Encryption Data
    112: {"type": "special_112"},                               # Additional Data (National Use)
    117: {"type": "special_117"},                               # Additional Transaction Reference Data 2
    118: {"type": "special_118"},                               # Additional Transaction Reference Data 3
    119: {"type": "special_119"},                               # Additional Data: Private Use 2
    120: {"type": "special_120"},                               # Record Data
    121: {"type": "llvar"},                                     # Authorizing Agent ID Code
    122: {"type": "special_122"},                               # Additional Private Data - National (dataset)
    123: {"type": "lllvar"},                                    # Receipt Free Text
    124: {"type": "special_124"},                               # Member-Defined Data
    125: {"type": "pin"},                                       # New PIN Data (8 bytes, masked)
    127: {"type": "lllvar"},                                    # Private Data (customer use)
}

# F43 (Card Acceptor Name/Location) - fixed 40 chars, positional subfields.
# Widths [22,1,13,1,3] confirmed against two independent samples with
# different merchant name/city content (bKash Limited/Dhaka/BGD and
# CHUNARUGHT FT C/HABIGANJ/BGD, and AROGGA LIMITED/DHAKA/BGD).
FIELD43_SUBFIELDS = [
    (1, "Card Acceptor Name", 22),
    (2, "Space Separator", 1),
    (3, "Card Acceptor City", 13),
    (4, "Space Separator", 1),
    (5, "Card Acceptor Country Code", 3),
]

# F90 (Original Data Elements) - fixed 42 chars, positional subfields.
# Widths [4,6,10,11,11] confirmed against a reversal (0400/0410) sample
# pair. Subfield 5's meaning is not yet confirmed - only observed as
# all-zero so far.
FIELD90_SUBFIELDS = [
    (1, "Original Message Type Indicator (MTI)", 4),
    (2, "Original Systems Trace Audit Number (STAN)", 6),
    (3, "Original Transmission Date and Time", 10),
    (4, "Original Amount, Transaction", 11),
    (5, "Reserved / unconfirmed", 11),
]

# F94 (Service Indicator) - used for AVS sign-on
FIELD94_SUBFIELDS = [
    (1, "Reserved for Future Use", 1),
    (2, "Acquirer/Issuer Indicator", 1),
    (3, "Address Data Indicator", 1),
]

# F48 (Additional Data - Private) subelement tag names, from the spec's
# "List of DE 48 subelements" table. Only the tags actually observed in
# confirmed sample traffic are listed with high confidence; others in the
# spec exist but haven't been exercised by a live sample yet.
FIELD48_TAG_NAMES = {
    "32": "Mastercard Assigned ID",
    "37": "Additional Merchant Data",
    "42": "Electronic Commerce Indicators",
    "43": "Universal Cardholder Authentication Field (UCAF)",
    "61": "POS Data Extended Condition Codes",
    "63": "Trace ID",
    "66": "Authentication Data",
    "71": "On-behalf Services",
    "87": "Card Validation Code Result",
    "92": "CVC 2",
}

# EMV tags seen in Field 55 (confirmed against live sample traffic).
EMV_TAG_NAMES = {
    "82": "Application Interchange Profile (AIP)",
    "84": "Dedicated File Name (AID)",
    "95": "Terminal Verification Results (TVR)",
    "9a": "Transaction Date",
    "9c": "Transaction Type",
    "5f2a": "Transaction Currency Code",
    "9f02": "Amount, Authorized (Numeric)",
    "9f03": "Amount, Other (Numeric)",
    "9f10": "Issuer Application Data (IAD)",
    "9f1a": "Terminal Country Code",
    "9f26": "Application Cryptogram (AC)",
    "9f27": "Cryptogram Information Data (CID)",
    "9f33": "Terminal Capabilities",
    "9f34": "Cardholder Verification Method (CVM) Results",
    "9f36": "Application Transaction Counter (ATC)",
    "9f37": "Unpredictable Number",
    "91": "Issuer Authentication Data",
}


# ---------------------------------------------------------------------------
# Field decoding dispatcher
# ---------------------------------------------------------------------------

def read_len_prefix(data: bytes, pos: int, nbytes: int):
    """Length prefix is ASCII/EBCDIC decimal DIGITS (not binary), width
    `nbytes` characters - confirmed 2-digit (LLVAR) and 3-digit (LLLVAR)
    variants across multiple fields and samples."""
    length = int(data[pos:pos + nbytes].decode("cp037"))
    return length, pos + nbytes


def decode_field(fnum: int, data: bytes, pos: int):
    """Decode one data field starting at pos. Returns (rendered_str, new_pos)."""
    fdef = FIELD_DEFS.get(fnum)
    if fdef is None:
        # Unknown field: assume 2-digit LLVAR EBCDIC text (safe default;
        # unconfirmed for any field not yet seen in sample traffic).
        length, pos2 = read_len_prefix(data, pos, nbytes=2)
        val = decode_ebcdic(data[pos2:pos2 + length])
        return val, pos2 + length

    t = fdef["type"]

    if t == "an":
        n = fdef["len"]
        val = decode_ebcdic(data[pos:pos + n])
        return val, pos + n

    if t == "llvar":
        length, pos2 = read_len_prefix(data, pos, nbytes=2)
        val = decode_ebcdic(data[pos2:pos2 + length])
        return val, pos2 + length

    if t == "lllvar":
        length, pos2 = read_len_prefix(data, pos, nbytes=3)
        val = decode_ebcdic(data[pos2:pos2 + length])
        return val, pos2 + length

    if t == "pin":
        # Field 52: fixed 8 bytes, always fully masked (encrypted PIN block)
        pos2 = pos + 8
        return "*" * 16, pos2

    if t == "special_3":
        return decode_field_3(data, pos)

    if t == "special_28":
        return decode_field_28(data, pos)

    if t == "special_43":
        return decode_field_43(data, pos)

    if t == "special_48":
        return decode_field_48(data, pos)

    if t == "special_54":
        return decode_field_54(data, pos)

    if t == "special_55":
        return decode_field_55(data, pos)

    if t == "special_90":
        return decode_field_90(data, pos)

    if t == "special_94":
        return decode_field_94(data, pos)

    if t == "special_95":
        return decode_field_95(data, pos)

    if t == "special_104":
        return decode_field_104(data, pos)

    if t == "special_105":
        return decode_field_105(data, pos)

    if t == "special_106":
        return decode_field_106(data, pos)

    if t == "special_108":
        return decode_field_108(data, pos)

    if t == "special_110":
        return decode_field_110(data, pos)

    if t == "special_112":
        return decode_field_112(data, pos)

    if t == "special_117":
        return decode_field_117(data, pos)

    if t == "special_118":
        return decode_field_118(data, pos)

    if t == "special_119":
        return decode_field_119(data, pos)

    if t == "special_120":
        return decode_field_120(data, pos)

    if t == "special_122":
        return decode_field_122(data, pos)

    if t == "special_124":
        return decode_field_124(data, pos)

    # fallback
    length, pos2 = read_len_prefix(data, pos, nbytes=2)
    val = decode_ebcdic(data[pos2:pos2 + length])
    return val, pos2 + length


FIELD3_SUBFIELDS = [
    (1, "Transaction Type", 2),
    (2, "Account Type, From", 2),
    (3, "Account Type, To", 2),
]


def decode_field_3(data: bytes, pos: int):
    n = 6
    body = decode_ebcdic(data[pos:pos + n])
    parts = []
    bpos = 0
    for sub, _name, width in FIELD3_SUBFIELDS:
        parts.append(f"F3.{sub})" + body[bpos:bpos + width])
        bpos += width
    return (SUBFIELD_SEP.join(parts) + SUBFIELD_SEP if parts else ""), pos + n


def decode_field_28(data: bytes, pos: int):
    """
    DE 28 (Amount, Transaction Fee): 9 positions
    - Position 1: Debit/Credit Indicator (C or D)
    - Positions 2-9: Amount (8 digits)
    """
    n = 9
    body = decode_ebcdic(data[pos:pos + n])
    indicator = body[0] if body else ""
    amount = body[1:] if len(body) > 1 else ""
    parts = [f"F28.1){indicator}", f"F28.2){amount}"]
    return (SUBFIELD_SEP.join(parts) + SUBFIELD_SEP if parts else ""), pos + n


def decode_field_43(data: bytes, pos: int):
    n = 40
    body = decode_ebcdic(data[pos:pos + n])
    parts = []
    bpos = 0
    for sub, _name, width in FIELD43_SUBFIELDS:
        parts.append(f"F43.{sub})" + body[bpos:bpos + width])
        bpos += width
    return (SUBFIELD_SEP.join(parts) + SUBFIELD_SEP if parts else ""), pos + n


def decode_field_90(data: bytes, pos: int):
    n = 42
    body = decode_ebcdic(data[pos:pos + n])
    parts = []
    bpos = 0
    for sub, _name, width in FIELD90_SUBFIELDS:
        parts.append(f"F90.{sub})" + body[bpos:bpos + width])
        bpos += width
    return (SUBFIELD_SEP.join(parts) + SUBFIELD_SEP if parts else ""), pos + n


def decode_field_94(data: bytes, pos: int):
    """
    DE 94 (Service Indicator): 7 positions
    - Subfield 1: Reserved for Future Use (1 char)
    - Subfield 2: Acquirer/Issuer Indicator (A, I, or B)
    - Subfield 3: Address Data Indicator (0-4)
    Positions 4-7 must contain spaces or zeros.
    """
    n = 7
    body = decode_ebcdic(data[pos:pos + n])
    parts = []
    for sub, _name, width in FIELD94_SUBFIELDS:
        start = (sub - 1) * width
        parts.append(f"F94.{sub})" + body[start:start + width])
    return (SUBFIELD_SEP.join(parts) + SUBFIELD_SEP if parts else ""), pos + n


def decode_field_95(data: bytes, pos: int):
    """
    DE 95 (Replacement Amounts): 42 positions
    - Subfield 1: Actual Amount, Transaction (12 digits)
    - Subfield 2: Actual Amount, Settlement (12 digits)
    - Subfield 3: Actual Amount, Cardholder Billing (12 digits)
    - Subfield 4: Zero Fill (6 digits)
    """
    n = 42
    body = decode_ebcdic(data[pos:pos + n])
    parts = [
        f"F95.1){body[0:12]}",
        f"F95.2){body[12:24]}",
        f"F95.3){body[24:36]}",
        f"F95.4){body[36:42]}",
    ]
    return (SUBFIELD_SEP.join(parts) + SUBFIELD_SEP if parts else ""), pos + n


def decode_field_48(data: bytes, pos: int):
    """
    DE 48 (Additional Data - Private): 3-digit LLLVAR outer length, then:
      - ONE untagged leading character - DE 48's "Transaction Category
        Code" (rendered here as subelement "01" for output consistency,
        though it's not actually a tagged subelement on the wire).
      - a repeating chain of tag(2 EBCDIC digits) + len(2 EBCDIC digits)
        + value, consumed until the outer length is exhausted.
    Confirmed against two independent samples (request carrying 2
    subelements, response carrying 6).
    """
    length, pos2 = read_len_prefix(data, pos, nbytes=3)
    body = decode_ebcdic(data[pos2:pos2 + length])
    parts = []
    bpos = 0
    if bpos < len(body):
        parts.append(("01", body[bpos]))
        bpos += 1
    while bpos < len(body):
        tag = body[bpos:bpos + 2]; bpos += 2
        if bpos + 2 > len(body):
            parts.append((tag, "<truncated>"))
            break
        sublen = int(body[bpos:bpos + 2]); bpos += 2
        val = body[bpos:bpos + sublen]; bpos += sublen
        parts.append((tag, val))
    
    # Sort parts by tag numeric value (01, 22, 37, 42, 43, 61, 66, 71, 75, 92)
    sorted_parts = sorted(parts, key=lambda x: int(x[0]) if x[0].isdigit() else 999)
    rendered_parts = [f"F48.{t})" + v for t, v in sorted_parts]
    return (SUBFIELD_SEP.join(rendered_parts) + SUBFIELD_SEP if rendered_parts else ""), pos2 + length


def decode_field_54(data: bytes, pos: int):
    """
    DE 54 (Additional Amounts): LLLVAR with repeating 20-byte occurrences.
    Each occurrence has:
    - Subfield 1: Account Type (2 digits)
    - Subfield 2: Amount Type (2 digits)
    - Subfield 3: Currency Code (3 digits)
    - Subfield 4: Debit or Credit Indicator (1 char, C or D)
    - Subfield 5: Amount (12 digits)
    """
    length, pos2 = read_len_prefix(data, pos, nbytes=3)
    body = decode_ebcdic(data[pos2:pos2 + length])
    parts = []
    bpos = 0
    occurrence = 1
    while bpos + 20 <= len(body):
        acct_type = body[bpos:bpos+2]; bpos += 2
        amt_type = body[bpos:bpos+2]; bpos += 2
        currency = body[bpos:bpos+3]; bpos += 3
        indicator = body[bpos:bpos+1]; bpos += 1
        amount = body[bpos:bpos+12]; bpos += 12
        parts.append(f"F54.{occurrence}.1){acct_type}")
        parts.append(f"F54.{occurrence}.2){amt_type}")
        parts.append(f"F54.{occurrence}.3){currency}")
        parts.append(f"F54.{occurrence}.4){indicator}")
        parts.append(f"F54.{occurrence}.5){amount}")
        occurrence += 1
    if bpos < len(body):
        parts.append(f"F54.{occurrence}){body[bpos:]}")
    return (SUBFIELD_SEP.join(parts) + SUBFIELD_SEP if parts else ""), pos2 + length


def decode_field_55(data: bytes, pos: int):
    """DE 55 (ICC/EMV data): 3-digit LLLVAR length, then a standard binary
    BER-TLV chain (real binary tags, not EBCDIC-encoded).

    NOTE on trailing separator: confirmed against two samples that F55's
    rendering only appends a trailing SUBFIELD_SEP when there is MORE
    THAN ONE tag - a single-tag F55 (e.g. a response carrying only tag
    91) renders with no trailing separator, while a multi-tag F55 does.
    This differs from every other composite field in this parser (F43/
    F48/F90/F122 all always append a trailing separator regardless of
    count) - do not "normalize" this without new sample evidence, it was
    verified byte-for-byte against two independent messages.
    """
    length, pos2 = read_len_prefix(data, pos, nbytes=3)
    body = data[pos2:pos2 + length]
    tlvs = decode_plain_tlv(body, value_renderer=render_f55_value)
    parts = [f"F55.{tag})" + val for tag, val in tlvs.items()]
    if not parts:
        rendered = ""
    elif len(parts) == 1:
        rendered = parts[0]
    else:
        rendered = SUBFIELD_SEP.join(parts) + SUBFIELD_SEP
    return rendered, pos2 + length


def decode_field_104(data: bytes, pos: int):
    """
    DE 104 (Digital Payment Data): LLLVAR with TLV subelements.
    Subelements:
    - 001: Digital Payment Cryptogram (28 chars, base64)
    - 002: Estimated Amount (12 digits)
    - 003: Remote Commerce Acceptor Identifier
    - 004: Digital Service Provider
    - 005: Digital Authentication Data
    """
    length, pos2 = read_len_prefix(data, pos, nbytes=3)
    body = decode_ebcdic(data[pos2:pos2 + length])
    parts = []
    bpos = 0
    while bpos < len(body):
        if bpos + 6 > len(body):
            parts.append((body[bpos:], "<truncated>"))
            break
        tag = body[bpos:bpos + 3]; bpos += 3
        sublen = int(body[bpos:bpos + 3]); bpos += 3
        val = body[bpos:bpos + sublen]; bpos += sublen
        parts.append((tag, val))
    rendered_parts = [f"F104.{t})" + v for t, v in parts]
    return (SUBFIELD_SEP.join(rendered_parts) + SUBFIELD_SEP if rendered_parts else ""), pos2 + length


def decode_field_105(data: bytes, pos: int):
    """
    DE 105 (Multi-Use Transaction Identification Data): 3-digit LLLVAR outer length,
    then a flat TLV chain of subelements (tag=3 digits, len=3 digits, value).
    
    Structure:
    - Subelement 001: Transaction Link ID (TLID) - 22 character alphanumeric value
    - Subelement 002: Economically Related Transaction Link Identifier - 22 chars
    - Subelement 003: Lifecycle TLID Validation - subfields
    """
    length, pos2 = read_len_prefix(data, pos, nbytes=3)
    body = decode_ebcdic(data[pos2:pos2 + length])
    parts = []
    bpos = 0
    
    while bpos < len(body):
        if bpos + 6 > len(body):
            parts.append((body[bpos:], "<truncated>"))
            break
        tag = body[bpos:bpos + 3]
        bpos += 3
        sublen = int(body[bpos:bpos + 3])
        bpos += 3
        val = body[bpos:bpos + sublen]
        bpos += sublen
        parts.append((tag, val))
    
    rendered_parts = [f"F105.{t})" + v for t, v in parts]
    return (SUBFIELD_SEP.join(rendered_parts) + SUBFIELD_SEP if rendered_parts else ""), pos2 + length


def decode_field_106(data: bytes, pos: int):
    """
    DE 106 (Fleet Service Data): LLLVAR with TLV subelements.
    Used for fleet card transactions (fuel items, non-fuel items, prompted data).
    """
    length, pos2 = read_len_prefix(data, pos, nbytes=3)
    body = decode_ebcdic(data[pos2:pos2 + length])
    parts = []
    bpos = 0
    while bpos < len(body):
        if bpos + 6 > len(body):
            parts.append((body[bpos:], "<truncated>"))
            break
        tag = body[bpos:bpos + 3]; bpos += 3
        sublen = int(body[bpos:bpos + 3]); bpos += 3
        val = body[bpos:bpos + sublen]; bpos += sublen
        parts.append((tag, val))
    rendered_parts = [f"F106.{t})" + v for t, v in parts]
    return (SUBFIELD_SEP.join(rendered_parts) + SUBFIELD_SEP if rendered_parts else ""), pos2 + length


def decode_field_108(data: bytes, pos: int):
    """
    DE 108 (Additional Transaction Reference Data): LLLVAR with TLV subelements.
    Used for Mastercard Send, Funding, and Gaming transactions.
    Subelements: 01=Receiver/Recipient Data, 02=Sender Data, 03=Transaction Reference Data
    """
    length, pos2 = read_len_prefix(data, pos, nbytes=3)
    body = decode_ebcdic(data[pos2:pos2 + length])
    parts = []
    bpos = 0
    while bpos < len(body):
        if bpos + 6 > len(body):
            parts.append((body[bpos:], "<truncated>"))
            break
        tag = body[bpos:bpos + 2]; bpos += 2
        sublen = int(body[bpos:bpos + 3]); bpos += 3
        val = body[bpos:bpos + sublen]; bpos += sublen
        parts.append((tag, val))
    rendered_parts = [f"F108.{t})" + v for t, v in parts]
    return (SUBFIELD_SEP.join(rendered_parts) + SUBFIELD_SEP if rendered_parts else ""), pos2 + length


def decode_field_110(data: bytes, pos: int):
    """
    DE 110 (Additional Data-2 / Encryption Data): LLLVAR raw pass-through.
    Contains PIN encryption data or key management encryption data.
    """
    length, pos2 = read_len_prefix(data, pos, nbytes=3)
    val = decode_ebcdic(data[pos2:pos2 + length])
    return val, pos2 + length


def decode_field_112(data: bytes, pos: int):
    """
    DE 112 (Additional Data - National Use): LLLVAR with TLV subelements.
    Used for regional data (Brazil, Chile, Colombia, France, Greece, India, etc.)
    """
    length, pos2 = read_len_prefix(data, pos, nbytes=3)
    body = decode_ebcdic(data[pos2:pos2 + length])
    parts = []
    bpos = 0
    while bpos < len(body):
        if bpos + 6 > len(body):
            parts.append((body[bpos:], "<truncated>"))
            break
        tag = body[bpos:bpos + 3]; bpos += 3
        sublen = int(body[bpos:bpos + 3]); bpos += 3
        val = body[bpos:bpos + sublen]; bpos += sublen
        parts.append((tag, val))
    rendered_parts = [f"F112.{t})" + v for t, v in parts]
    return (SUBFIELD_SEP.join(rendered_parts) + SUBFIELD_SEP if rendered_parts else ""), pos2 + length


def decode_field_117(data: bytes, pos: int):
    """
    DE 117 (Additional Transaction Reference Data 2): LLLVAR with TLV subelements.
    Used for intermediary party data (Transaction Initiator, Financial Intermediary 1/2).
    """
    length, pos2 = read_len_prefix(data, pos, nbytes=3)
    body = decode_ebcdic(data[pos2:pos2 + length])
    parts = []
    bpos = 0
    while bpos < len(body):
        if bpos + 6 > len(body):
            parts.append((body[bpos:], "<truncated>"))
            break
        tag = body[bpos:bpos + 3]; bpos += 3
        sublen = int(body[bpos:bpos + 3]); bpos += 3
        val = body[bpos:bpos + sublen]; bpos += sublen
        parts.append((tag, val))
    rendered_parts = [f"F117.{t})" + v for t, v in parts]
    return (SUBFIELD_SEP.join(rendered_parts) + SUBFIELD_SEP if rendered_parts else ""), pos2 + length


def decode_field_118(data: bytes, pos: int):
    """
    DE 118 (Additional Transaction Reference Data 3): LLLVAR with TLV subelements.
    Used for Age Verification Request/Response.
    """
    length, pos2 = read_len_prefix(data, pos, nbytes=3)
    body = decode_ebcdic(data[pos2:pos2 + length])
    parts = []
    bpos = 0
    while bpos < len(body):
        if bpos + 6 > len(body):
            parts.append((body[bpos:], "<truncated>"))
            break
        tag = body[bpos:bpos + 3]; bpos += 3
        sublen = int(body[bpos:bpos + 3]); bpos += 3
        val = body[bpos:bpos + sublen]; bpos += sublen
        parts.append((tag, val))
    rendered_parts = [f"F118.{t})" + v for t, v in parts]
    return (SUBFIELD_SEP.join(rendered_parts) + SUBFIELD_SEP if rendered_parts else ""), pos2 + length


def decode_field_119(data: bytes, pos: int):
    """
    DE 119 (Additional Data: Private Use 2): LLLVAR with TLV subelements.
    Used for MDES service indicators (Clearing Indicator, Acquirer Reference ID).
    """
    length, pos2 = read_len_prefix(data, pos, nbytes=3)
    body = decode_ebcdic(data[pos2:pos2 + length])
    parts = []
    bpos = 0
    while bpos < len(body):
        if bpos + 6 > len(body):
            parts.append((body[bpos:], "<truncated>"))
            break
        tag = body[bpos:bpos + 3]; bpos += 3
        sublen = int(body[bpos:bpos + 3]); bpos += 3
        val = body[bpos:bpos + sublen]; bpos += sublen
        parts.append((tag, val))
    rendered_parts = [f"F119.{t})" + v for t, v in parts]
    return (SUBFIELD_SEP.join(rendered_parts) + SUBFIELD_SEP if rendered_parts else ""), pos2 + length


def decode_field_120(data: bytes, pos: int):
    """
    DE 120 (Record Data): LLLVAR raw pass-through.
    Used for AVS (Address Verification Service) data and file maintenance.
    """
    length, pos2 = read_len_prefix(data, pos, nbytes=3)
    val = decode_ebcdic(data[pos2:pos2 + length])
    return val, pos2 + length


def decode_field_122(data: bytes, pos: int):
    """
    DE 122 (Additional Private Data - National): 3-digit LLLVAR outer
    length, then a FLAT (non-recursive) repeating chain of tag(3 digits)
    + len(3 digits) + value. Confirmed against a sample where the single
    tag "001"'s value contains text that superficially resembles further
    tag/len/value triples - deliberately NOT decomposed further here,
    matching the reference tool's behavior this parser is validated
    against. Do not change this to a recursive parse without new sample
    evidence showing that's actually expected.
    """
    length, pos2 = read_len_prefix(data, pos, nbytes=3)
    body = decode_ebcdic(data[pos2:pos2 + length])
    parts = []
    bpos = 0
    while bpos < len(body):
        if bpos + 6 > len(body):
            parts.append((body[bpos:], "<truncated>"))
            break
        tag = body[bpos:bpos + 3]; bpos += 3
        sublen = int(body[bpos:bpos + 3]); bpos += 3
        val = body[bpos:bpos + sublen]; bpos += sublen
        parts.append((tag, val))
    rendered_parts = [f"F122.{t})" + v for t, v in parts]
    return (SUBFIELD_SEP.join(rendered_parts) + SUBFIELD_SEP if rendered_parts else ""), pos2 + length


def decode_field_124(data: bytes, pos: int):
    """
    DE 124 (Member-Defined Data): LLLVAR raw pass-through.
    Used for MDES tokenization, Mastercard Send, and Brazil Maestro data.
    """
    length, pos2 = read_len_prefix(data, pos, nbytes=3)
    val = decode_ebcdic(data[pos2:pos2 + length])
    return val, pos2 + length


# ---------------------------------------------------------------------------
# Description dictionaries
# ---------------------------------------------------------------------------

HEADER_DESCRIPTIONS = {
    "MTI": "Message Type Indicator",
}

FIELD_DESCRIPTIONS = {
    2: "Primary Account Number (PAN)",
    3: "Processing Code",
    4: "Amount, Transaction",
    5: "Amount, Reconciliation",
    6: "Amount, Cardholder Billing",
    7: "Transmission Date and Time",
    9: "Conversion Rate, Reconciliation",
    10: "Conversion Rate, Cardholder Billing",
    11: "Systems Trace Audit Number (STAN)",
    12: "Time, Local Transaction",
    13: "Date, Local Transaction",
    14: "Date, Expiration",
    15: "Date, Settlement",
    16: "Date, Conversion",
    18: "Merchant Category Code (MCC)",
    19: "Acquiring Institution Country Code",
    22: "Point of Service (POS) Entry Mode",
    23: "Card Sequence Number",
    26: "POS PIN Capture Code",
    28: "Amount, Transaction Fee",
    32: "Acquiring Institution ID Code",
    33: "Forwarding Institution ID Code",
    35: "Track 2 Data",
    37: "Retrieval Reference Number (RRN)",
    38: "Authorization ID Response",
    39: "Response Code",
    41: "Card Acceptor Terminal ID",
    42: "Card Acceptor ID Code",
    43: "Card Acceptor Name/Location",
    44: "Additional Response Data",
    48: "Additional Data - Private Use",
    49: "Currency Code, Transaction",
    50: "Currency Code, Reconciliation",
    51: "Currency Code, Cardholder Billing",
    52: "PIN Data",
    54: "Additional Amounts",
    55: "ICC System Related Data (EMV)",
    56: "Payment Account Data (PAR)",
    61: "Point-of-Service (POS) Data",
    62: "Intermediate Network Facility (INF) Data",
    63: "Private / Network-Assigned Reference",
    67: "Extended Payment Code",
    90: "Original Data Elements",
    94: "Service Indicator",
    95: "Replacement Amounts",
    96: "Message Security Code",
    102: "Account ID 1",
    103: "Account ID 2",
    104: "Digital Payment Data",
    105: "Multi-Use Transaction Identification Data",
    106: "Fleet Service Data",
    108: "Additional Transaction Reference Data",
    110: "Additional Data-2 / Encryption Data",
    112: "Additional Data (National Use)",
    117: "Additional Transaction Reference Data 2",
    118: "Additional Transaction Reference Data 3",
    119: "Additional Data: Private Use 2",
    120: "Record Data",
    121: "Authorizing Agent ID Code",
    122: "Additional Private Use Data - National",
    123: "Receipt Free Text",
    124: "Member-Defined Data",
    125: "New PIN Data",
    127: "Private Data",
}

FIELD3_DESCRIPTIONS = {sub: name for sub, name, _ in FIELD3_SUBFIELDS}
FIELD43_DESCRIPTIONS = {sub: name for sub, name, _ in FIELD43_SUBFIELDS}
FIELD90_DESCRIPTIONS = {sub: name for sub, name, _ in FIELD90_SUBFIELDS}
FIELD94_DESCRIPTIONS = {sub: name for sub, name, _ in FIELD94_SUBFIELDS}


def describe_label(label: str) -> str:
    """
    Resolve a human-readable description for a compact-format label such
    as 'MTI', 'F2', 'F43.1', 'F48.32', 'F55.9f26', 'F90.4', 'F122.001'.
    """
    if label == "MTI":
        return HEADER_DESCRIPTIONS["MTI"]

    parts = label[1:].split(".")
    try:
        fnum = int(parts[0])
    except ValueError:
        return "Unknown field"

    base_desc = FIELD_DESCRIPTIONS.get(fnum, "Unrecognized / proprietary field")

    if len(parts) == 1:
        return base_desc

    if fnum == 3:
        sub = int(parts[1])
        return f"{base_desc} - Subfield {sub}: {FIELD3_DESCRIPTIONS.get(sub, 'Reserved')}"

    if fnum == 28:
        sub = int(parts[1])
        names = {1: "Debit/Credit Indicator", 2: "Amount"}
        return f"{base_desc} - Subfield {sub}: {names.get(sub, 'Reserved')}"

    if fnum == 43:
        sub = int(parts[1])
        return f"{base_desc} - Subfield {sub}: {FIELD43_DESCRIPTIONS.get(sub, 'Reserved')}"

    if fnum == 54:
        sub_parts = parts[1].split(".")
        if len(sub_parts) == 2:
            occ = sub_parts[0]
            sf = sub_parts[1]
            names = {1: "Account Type", 2: "Amount Type", 3: "Currency Code", 4: "Debit/Credit Indicator", 5: "Amount"}
            return f"{base_desc} - Occurrence {occ}, Subfield {sf}: {names.get(int(sf), 'Reserved')}"
        return f"{base_desc} - Occurrence {parts[1]}"

    if fnum == 90:
        sub = int(parts[1])
        return f"{base_desc} - Subfield {sub}: {FIELD90_DESCRIPTIONS.get(sub, 'Reserved')}"

    if fnum == 94:
        sub = int(parts[1])
        return f"{base_desc} - Subfield {sub}: {FIELD94_DESCRIPTIONS.get(sub, 'Reserved')}"

    if fnum == 95:
        sub = int(parts[1])
        names = {1: "Actual Amount, Transaction", 2: "Actual Amount, Settlement", 3: "Actual Amount, Cardholder Billing", 4: "Zero Fill"}
        return f"{base_desc} - Subfield {sub}: {names.get(sub, 'Reserved')}"

    if fnum == 48:
        tag = parts[1]
        if tag == "01":
            return f"{base_desc} - Transaction Category Code (leading untagged subelement)"
        name = FIELD48_TAG_NAMES.get(tag, f"Subelement {tag} (not yet confirmed against a live sample)")
        return f"{base_desc} - Subelement {tag}: {name}"

    if fnum == 55:
        tag = parts[1].lower()
        return f"{base_desc} - {EMV_TAG_NAMES.get(tag, 'Proprietary/unrecognized EMV tag ' + tag.upper())}"

    if fnum == 104:
        tag = parts[1]
        names = {"001": "Digital Payment Cryptogram", "002": "Estimated Amount", 
                 "003": "Remote Commerce Acceptor Identifier", "004": "Digital Service Provider",
                 "005": "Digital Authentication Data"}
        return f"{base_desc} - Subelement {tag}: {names.get(tag, 'Unrecognized')}"

    if fnum == 105:
        tag = parts[1]
        names = {"001": "Transaction Link ID (TLID)", "002": "Economically Related TLID", 
                 "003": "Lifecycle TLID Validation"}
        return f"{base_desc} - Subelement {tag}: {names.get(tag, 'Unrecognized')}"

    if fnum == 106:
        tag = parts[1]
        names = {"001": "Fleet Prompted Data", "002": "Merchant Fleet Spend Control Capability",
                 "003": "Fleet Spend Control Override Items", "004": "Fleet Fuel Information",
                 "005": "Fleet Non-Fuel Information"}
        return f"{base_desc} - Subelement {tag}: {names.get(tag, 'Unrecognized')}"

    if fnum == 112:
        tag = parts[1]
        return f"{base_desc} - Subelement {tag} (Regional/National Use)"

    if fnum == 122:
        tag = parts[1]
        if tag == "001":
            return f"{base_desc} - Subelement 001 (Merchant/Additional Data; internal structure not decomposed - see module docstring)"
        return f"{base_desc} - Subelement {tag} (not yet confirmed against a live sample)"

    return base_desc


# ---------------------------------------------------------------------------
# Top level message parser
# ---------------------------------------------------------------------------

def _split_compact_field(fnum: int, rendered: str, raw_hex: str):
    """
    Given the compact rendering of one top-level field (which may contain
    several 'Label)value' groups terminated by SUBFIELD_SEP for composite
    fields), return a list of (label, value) row tuples.
    """
    rows = []
    if fnum in (3, 28, 43, 48, 54, 55, 90, 94, 95, 104, 105, 106, 108, 112, 117, 118, 119, 122):
        groups = [g for g in rendered.split(SUBFIELD_SEP) if g]
        for g in groups:
            if ")" in g:
                label, value = g.split(")", 1)
            else:
                label, value = f"F{fnum}", g
            rows.append((label, value))
        if not rows:
            rows.append((f"F{fnum}", ""))
    else:
        rows.append((f"F{fnum}", rendered))
    return rows


def parse_message_full(raw_hex: str, debug: bool = False):
    """
    Full parse returning both the compact string (identical to
    parse_message's output) and a list of detailed rows for reporting:
    each row is (label, description, raw_hex, value).
    """
    data = clean_hex(raw_hex)
    header, pos = parse_header(data)
    mti_start = 0
    present_fields, pos = parse_all_bitmaps(data, pos)
    if debug:
        print(f"[debug] mti={header['MTI']} fields_start_pos={pos}", file=sys.stderr)

    rows = [("MTI", describe_label("MTI"), data[mti_start:mti_start + 4].hex().upper(), header["MTI"])]
    out_parts = [f"F{fnum}" for fnum in []]  # placeholder, built below

    out_parts = []
    for fnum in sorted(present_fields):
        start = pos
        try:
            rendered, pos = decode_field(fnum, data, pos)
        except Exception as e:
            rendered = f"<parse-error: {e}>"
        if debug:
            print(f"[debug] F{fnum}: bytes[{start}:{pos}] = {data[start:pos].hex()} -> {rendered!r}", file=sys.stderr)

        raw_hex_field = data[start:pos].hex().upper()
        for label, value in _split_compact_field(fnum, rendered, raw_hex_field):
            rows.append((label, describe_label(label), raw_hex_field, value))

        if fnum in (3, 28, 43, 48, 54, 55, 90, 94, 95, 104, 105, 106, 108, 112, 117, 118, 119, 122):
            out_parts.append(rendered)  # already contains F{n}.{sub}) formatting
        else:
            out_parts.append(f"F{fnum})" + rendered)

    mti = header["MTI"]
    compact = f"{mti[1:] if mti.startswith('0') else mti}: " + FIELD_SEP.join(out_parts) + FIELD_SEP
    return compact, rows


def parse_message(raw_hex: str, debug: bool = False) -> str:
    return parse_message_full(raw_hex, debug=debug)[0]


# ---------------------------------------------------------------------------
# Report export: colored console, .txt, .xlsx (mirrors the Visa parser's
# TIC-style output convention: one folder per message named
# <RRN_or_STAN>_<MTI>, containing a .txt and a .xlsx report, plus a
# colored summary printed to console)
# ---------------------------------------------------------------------------

class _Ansi:
    RESET = "\033[0m"
    BOLD = "\033[1m"
    CYAN = "\033[96m"
    YELLOW = "\033[93m"
    GREEN = "\033[92m"
    MAGENTA = "\033[95m"
    RED = "\033[91m"
    DIM = "\033[2m"


def _supports_color() -> bool:
    if sys.platform == "win32":
        try:
            import colorama
            colorama.just_fix_windows_console()
            return True
        except Exception:
            return False
    return sys.stdout.isatty()


def print_colored_report(compact: str, rows, mti: str):
    use_color = _supports_color()

    def c(text, color):
        return f"{color}{text}{_Ansi.RESET}" if use_color else text

    print()
    print(c(f"===== Mastercard ISO8583 Message (MTI {mti}) =====", _Ansi.BOLD + _Ansi.CYAN))
    for label, desc, raw_hex, value in rows:
        label_c = c(f"{label:<14}", _Ansi.YELLOW)
        desc_c = c(f"{desc:<55}", _Ansi.DIM)
        value_c = c(value, _Ansi.GREEN)
        print(f"  {label_c} {desc_c} {value_c}")
    print()


def get_report_name(rows, mti: str) -> str:
    """Folder/file base name: <RRN_or_STAN>_<MTI>."""
    row_map = {label: value for label, _desc, _raw, value in rows}
    rrn = row_map.get("F37", "").strip()
    stan = row_map.get("F11", "").strip()
    key = rrn if rrn else (stan if stan else "UNKNOWN")
    return f"{key}_{mti}"


def write_txt_report(path: str, raw_hex: str, compact: str, rows):
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"Generated: {datetime.now().isoformat(timespec='seconds')}\n\n")

        f.write("=" * 120 + "\n")
        f.write("RAW MESSAGE (EXACT INPUT)\n")
        f.write("=" * 120 + "\n")
        f.write(raw_hex)
        f.write("\n\n")

        f.write("=" * 120 + "\n")
        f.write("COMPACT PARSED MESSAGE\n")
        f.write("=" * 120 + "\n")
        f.write(compact)
        f.write("\n\n")

        f.write("=" * 120 + "\n")
        f.write("FIELD-BY-FIELD BREAKDOWN\n")
        f.write("=" * 120 + "\n")

        field_width = max(len("Field"), max(len(str(label)) for label, _, _, _ in rows)) + 2
        desc_width = max(len("Description"), max(len(str(desc)) for _, desc, _, _ in rows)) + 2

        header = f"{'Field':<{field_width}}{'Description':<{desc_width}}{'Value'}"
        f.write(header + "\n")
        f.write("-" * len(header) + "\n")

        for label, desc, raw, value in rows:
            value = "" if value is None else str(value)
            f.write(f"{label:<{field_width}}{desc:<{desc_width}}{value}\n")


def write_detail_report(path: str, rows):
    with open(path, "w", encoding="utf-8") as f:
        f.write(format_field_report(rows))


def write_xlsx_report(path: str, raw_hex: str, compact: str, rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl is not installed - skipping .xlsx export "
              "(install with: pip install openpyxl --break-system-packages)")
        return

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "Raw Message (hex)"
    ws_summary["A1"].font = Font(bold=True)
    ws_summary["A2"] = raw_hex
    ws_summary["A2"].alignment = Alignment(wrap_text=True)
    ws_summary["A4"] = "Compact Parsed Message"
    ws_summary["A4"].font = Font(bold=True)
    ws_summary["A5"] = compact
    ws_summary["A5"].alignment = Alignment(wrap_text=True)
    ws_summary.column_dimensions["A"].width = 120

    ws = wb.create_sheet("Fields")
    headers = ["Field", "Description", "Raw Hex", "Value"]
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for r, (label, desc, raw, value) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=raw)
        ws.cell(row=r, column=4, value=value)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 40
    ws.auto_filter.ref = f"A1:D{len(rows) + 1}"

    wb.save(path)


def export_report(raw_hex_input: str, out_base_dir: str = r"D:\MASTERCARD_Output", debug: bool = False):
    import os

    compact, rows = parse_message_full(raw_hex_input, debug=debug)
    mti = compact.split(":", 1)[0].strip()
    name = get_report_name(rows, mti)

    folder = os.path.join(out_base_dir, name)
    os.makedirs(folder, exist_ok=True)

    txt_path = os.path.join(folder, f"{name}.txt")
    xlsx_path = os.path.join(folder, f"{name}.xlsx")
    detail_path = os.path.join(folder, f"{name}_details.txt")

    write_txt_report(txt_path, raw_hex_input.strip(), compact, rows)
    write_xlsx_report(xlsx_path, raw_hex_input.strip(), compact, rows)
    write_detail_report(detail_path, rows)

    print_colored_report(compact, rows, mti)
    print(f"Saved: {txt_path}")
    print(f"Saved: {xlsx_path}")
    print(f"Saved: {detail_path}")

    return compact, folder


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    debug = "--debug" in sys.argv
    no_export = "--no-export" in sys.argv
    out_dir = r"D:\MASTERCARD_Output"
    for a in sys.argv[1:]:
        if a.startswith("--outdir="):
            out_dir = a.split("=", 1)[1]

    if args:
        with open(args[0], "r") as f:
            raw = f.read()
    else:
        print("Paste the raw ISO8583 hex message below.", flush=True)
        print("Press Enter on an empty line when you're done:\n", flush=True)
        lines = []
        try:
            while True:
                line = input()
                if line.strip() == "":
                    break
                lines.append(line)
        except EOFError:
            pass
        raw = "\n".join(lines)

    if not raw.strip():
        print("\nNo input received - nothing to parse.")
        sys.exit(1)

    try:
        clean_hex(raw)
    except Exception as e:
        print(f"\nCould not parse input as hex: {e}")
        sys.exit(1)

    try:
        if no_export:
            print("\n=== Parsed message ===")
            print(parse_message(raw, debug=debug))
        else:
            export_report(raw, out_base_dir=out_dir, debug=debug)
    except Exception as e:
        print(f"<parser error: {e}>")